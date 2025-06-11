import openpyxl
from openpyxl.styles import PatternFill
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Open a file dialog to select the Excel file
Tk().withdraw()  # Hide the root window
file_path = askopenfilename(title="Select the Mongoose Contact List Excel file", filetypes=[("Excel files", "*.xlsx")])

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Create a new sheet for rows where sms_allowed is equal to 0
no_sms_sheet = wb.create_sheet(title="NO SMS")

# Define colors for highlighting cells
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# Copy header row to NO SMS sheet
header = [cell.value for cell in sheet[1]]
no_sms_sheet.append(header)

# Process each row
for row in list(sheet.iter_rows(min_row=2)):
    sms_allowed = row[18].value  # Column S

    if sms_allowed == 0:
        no_sms_sheet.append([cell.value for cell in row])
        sheet.delete_rows(row[0].row)
        continue

    # Compare columns C, D, E after stripping whitespace
    names = [str(row[i].value).strip() for i in [2, 3, 4]]
    c_val, d_val, e_val = names

    # If C is blank and D == E (not blank), auto-fill all with D
    if c_val == ("" or "None") and d_val == e_val and d_val != "":
        for i in [2, 3, 4]:
            row[i].value = row[3].value  # Use D's value
    elif len(set(names)) > 1:
        # Highlight the cells in red
        for i in [2, 3, 4]:
            row[i].fill = red_fill
        # Ask user which value to use
        print(f"\nRow {row[0].row}: The following values in columns C, D, E are different:")
        for idx, i in enumerate([2, 3, 4]):
            print(f"{idx + 1}: {row[i].value}")
        choice = None
        while choice not in [1, 2, 3]:
            try:
                choice = int(input("Which value do you want to use for all three? (1: Admissions/2: OTP/3: Legal): "))
            except ValueError:
                continue
        selected_value = row[[2, 3, 4][choice - 1]].value
        for i in [2, 3, 4]:
            row[i].value = selected_value
            #If #2 was chosen, remove the red fill, no changes need to be made in OTP Admin Console
            if choice == 2:
                row[i].fill = PatternFill()

    # Text replace appl_qtr (column L)
    qtr_map = {1: "Winter", 2: "Spring", 3: "Summer", 4: "Autumn"}
    qtr_cell = row[11]
    if qtr_cell.value in qtr_map:
        qtr_cell.value = qtr_map[qtr_cell.value]

    # Text replace appl_type (column M)
    # Mapping application types to their descriptions.
    # Both "2" and "4" map to "Transfer" because they represent similar categories.
    type_map = {"1": "First Year", "2": "Transfer", "4": "Transfer", "5": "Post Bac"}
    type_cell = row[12]
    if str(type_cell.value) in type_map:
        type_cell.value = type_map[type_cell.value]

    # Validate phone numbers in columns H, I, J
    for i in [7, 8, 9]:
        phone = str(row[i].value)
        if not (phone.isdigit() and len(phone) == 10):
            row[i].fill = orange_fill
#Closing the loop for processing rows

# Delete columns in reverse order to avoid index shifting
for col_letter in ['T', 'S', 'J', 'I', 'G', 'E', 'C', 'A']:
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    sheet.delete_cols(col_idx)

#Insert new blank columns, must be done right to left to avoid shifting issues
# Insert a new column between I and J 
sheet.insert_cols(10)
# Insert a new column between D and E 
sheet.insert_cols(5) 

# Fill in all the cells in column E with "Bothell"
for row in range(2, sheet.max_row + 1):
    sheet.cell(row=row, column=5).value = "Bothell"

# Fill in all the cells in column K with "2025"
for row in range(2, sheet.max_row + 1):
    sheet.cell(row=row, column=11).value = "2025"


#Replace the first cell of each column A through N with new text
for col in range(1, 15):
    cell = sheet.cell(row=1, column=col)
    if col == 1:
        cell.value = "ContactId"
    elif col == 2:
        cell.value = "FirstName"
    elif col == 3:
        cell.value = "LastName"
    elif col == 4:
        cell.value = "MobileNumber"
    elif col == 5:
        cell.value = "Application_Campus"
    elif col == 6:
        cell.value = "Application_Year"
    elif col == 7:
        cell.value = "Application_Quarter"
    elif col == 8:
        cell.value = "Application_Type_Class"
    elif col == 9:
        cell.value = "SessionCode"
    elif col == 10:
        cell.value = "Student_Number"
    elif col == 11:
        cell.value = "Import_Year"
    elif col == 12:
        cell.value = "CoachID"
    elif col == 13:
        cell.value = "Coach First Name"
    elif col == 14:
        cell.value = "Session_Date"

# Save the updated workbook
output_path = file_path.replace(".xlsx", "_Upload.xlsx")
wb.save(output_path)
print(f"Updated file saved as: {output_path}")
